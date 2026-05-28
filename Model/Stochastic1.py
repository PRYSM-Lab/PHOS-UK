#======================================Input Data========================================================*
# %% Reading Data from Excel-----------------------
import pandas as pd
import numpy as np
from numpy import unravel_index
from sklearn.cluster import KMeans
from collections import Counter, defaultdict
from scipy.stats import gaussian_kde
from scipy.interpolate import interp1d
import tkinter as tk
from tkinter import simpledialog, messagebox
import subprocess
import concurrent.futures
from tkinter import simpledialog
import threading
from tkinter import filedialog
import openpyxl.utils.cell as cell_util

import os



root = tk.Tk()
root.withdraw() 

file_path1 = filedialog.askopenfilename(
    title="Select Excel File",
    filetypes=[("Excel Files", "*.xlsx *.xls")] 
)

if file_path1:
    excel_data = pd.ExcelFile(file_path1, engine='openpyxl')


import tkinter as tk
from tkinter import simpledialog, messagebox
import subprocess
import threading
import pandas as pd



global n_slice  

def run_intraday(n_slice_input):
    print(f"Running scenario.py with time-slice: {n_slice_input}")

    result = subprocess.run(
        ['python', 'scenario.py', str(n_slice_input)],  
        capture_output=True,
        text=True
    )

    if result.returncode == 0:  
        print("Duration of each time slice obtained.")  
    else:
        print(f"Error occurred: {result.stderr}")

def get_time_slice():
    global n_slice  
    root = tk.Tk()
    root.withdraw()  
    root.geometry("800x400")
    n_slice = simpledialog.askinteger("Input", "Enter number of time-slices:\n(Max value: 24)")
    
    print(f"User entered n_slice: {n_slice}")  

    if n_slice is None:
        print("Error: No value entered for n_slice!")
        return  

    thread = threading.Thread(target=run_intraday, args=(n_slice,), daemon=True)
    thread.start()
    thread.join()  

get_time_slice()


file_path3 = os.path.join(os.getcwd(), 'Time Slice.xlsx')
#excel_data = pd.ExcelFile(file_path1, engine='openpyxl')
excel_data3 = pd.ExcelFile(file_path3, engine='openpyxl')


df_theta=excel_data3.parse('theta', header=None, usecols="A:C", skiprows=1, nrows=121)
df_Hsum=excel_data3.parse('Hsum_index', header=None, usecols="A:C", skiprows=1, nrows=121)



import tkinter as tk
from tkinter import simpledialog, messagebox
import subprocess
import threading
import pandas as pd



global num_scenarios  

def run_scenario_reduction(num_scenarios_input):
    print(f"Running scenred1.py with scenario number: {num_scenarios_input}")

    result = subprocess.run(
        ['python', 'scenred1.py', str(num_scenarios_input)],  
        capture_output=True,
        text=True
    )

    if result.returncode == 0:  
        print("ScenRed finds the reduced scenarios and associated probabilities.")  
    else:
        print(f"Error occurred: {result.stderr}")

def get_number_scenario():
    global num_scenarios 
    root = tk.Tk()
    root.withdraw()  
    root.geometry("800x400")
    num_scenarios = simpledialog.askinteger("Input", "Enter number of scenario:\n(Max value: 1125)")
    
    print(f"User entered num_scenarios: {num_scenarios}")  

    if num_scenarios is None:
        print("Error: No value entered for num_scenarios!")
        return  

    thread = threading.Thread(target=run_scenario_reduction, args=(num_scenarios,), daemon=True)
    thread.start()
    thread.join()  

get_number_scenario()


file_path4 = os.path.join(os.getcwd(), 'ReducedScenarios.xlsx')
#excel_data = pd.ExcelFile(file_path1, engine='openpyxl')
excel_data4 = pd.ExcelFile(file_path4, engine='openpyxl')

num_columns1 = num_scenarios * 13+2
end_col_letter1 = cell_util.get_column_letter(num_columns1)

usecols1 = f"A:{end_col_letter1}"  



num_columns3 = num_scenarios * 13+3
end_col_letter3 = cell_util.get_column_letter(num_columns1)

usecols3 = f"A:{end_col_letter3}"

num_columns2 = num_scenarios+1 
end_col_letter2 = cell_util.get_column_letter(num_columns2)

usecols2 = f"A:{end_col_letter2}"  


num_columns4 = num_scenarios+2 
end_col_letter4 = cell_util.get_column_letter(num_columns4)

usecols4 = f"A:{end_col_letter4}" 

#file_path1 = r'C:\Users\Mohammed\newhydro_clusters.xlsx'
#file_path2 = r'C:\Users\Mohammed\stoch_input.xlsx'
file_path2 = os.path.join(os.getcwd(), 'stoch_input.xlsx')
#excel_data = pd.ExcelFile(file_path1, engine='openpyxl')
excel_data2 = pd.ExcelFile(file_path2, engine='openpyxl')


Regions_data = excel_data.parse('Regions')
Sets_data = excel_data.parse('Sets')
Distances_data = excel_data.parse('Distances')
General_data = excel_data.parse('General')
Heating_data = excel_data.parse('Heating')
Storage_data = excel_data.parse('Storage')
Production_data = excel_data.parse('Production')
Renewables_data = excel_data.parse('Renewables')
Emissions_data = excel_data.parse('Emissions')
H2Pipline_data = excel_data.parse('H2Pipeline')
CO2Pipline_data = excel_data.parse('CO2Pipeline')
CO2Reservior_data = excel_data.parse('CO2Reservoir')
df_Biomass = excel_data.parse('Biomass', header=None, usecols="B:C", skiprows=3, nrows=13)
df_bio = excel_data.parse('General', header=None, usecols="E:F", skiprows=47, nrows=5)
df_dc = excel_data.parse('General', header=None, usecols="A:B", skiprows=57, nrows=5)
df_cccH = excel_data.parse('H2Pipeline', header=None, usecols="A:B", skiprows=32, nrows=3)
df_cccC_Onshore = excel_data.parse('CO2Pipeline', header=None, usecols="A:B", skiprows=59, nrows=2)
df_cccC_offshore = excel_data.parse('CO2Pipeline', header=None, usecols="D:E", skiprows=59, nrows=2)
df_ct = excel_data.parse('CO2Pipeline', header=None, usecols="B:F", skiprows=2, nrows=2)
df_Cstart = excel_data.parse('Production', header=None, usecols="A:B", skiprows=61, nrows=4)
df_Cshut = excel_data.parse('Production', header=None, usecols="C:D", skiprows=61, nrows=4)
df_DT = excel_data.parse('Production', header=None, usecols="C:D", skiprows=51, nrows=4)
df_ec = excel_data.parse('Renewables', header=None, usecols="D:G", skiprows=17, nrows=2)
df_emtarget = excel_data.parse('Emissions', header=None, usecols="C:G", skiprows=39, nrows=2)
df_DistRes = excel_data.parse('Distances', header=None, usecols="B:D", skiprows=23, nrows=3)
df_DistSt = excel_data.parse('Distances', header=None, usecols="B:D", skiprows=31, nrows=4)
df_Dist = excel_data.parse('Distances', header=None, usecols="B:N", skiprows=4, nrows=13)
df_DistPipe = excel_data.parse('Distances', header=None, usecols="R:AD", skiprows=38, nrows=13)


pccostWE_data=excel_data4.parse('pccost')
pocostFWE_data=excel_data4.parse('pccostF')
pccosVWE_data=excel_data4.parse('pocostV')
rccost_data=excel_data4.parse('rccost')
Biomass_data = excel_data4.parse('biomass')
df_Bio= excel_data4.parse('biomass', header=None, usecols=usecols2, skiprows=1, nrows=3)
gasprice_data=excel_data4.parse('Prices')
etaWE_data=excel_data4.parse('eta')
df_eta= excel_data4.parse('eta', header=None, usecols=usecols2, skiprows=1, nrows=3)
Cluster_data= excel_data2.parse('Cluster')
df_cgas = excel_data4.parse('Prices', header=None, usecols=usecols2, skiprows=1, nrows=3)
#df_Hsum = excel_data2.parse('Hsum', header =None, usecols="A:G",skiprows=2, nrows=46)
df_ComDem= excel_data4.parse('Commercial', header=None, usecols=usecols1, skiprows=1, nrows=120)
df_DomDem= excel_data4.parse('Domestic', header=None, usecols=usecols1, skiprows=1, nrows=120)
df_IndDem= excel_data4.parse('Industrial', header=None, usecols=usecols1, skiprows=1, nrows=120)
df_hdem=excel_data4.parse('Dem', header=None, usecols=usecols4, skiprows=1, nrows=12)
df_Availability = excel_data4.parse('Renewable', header=None, usecols=usecols3, skiprows=1, nrows=360)
df_rccost=excel_data4.parse('rccost',header=None, usecols=usecols4, skiprows=1, nrows=9 )
Prob_data=excel_data4.parse('Probabilities')

# %% Building the Model===================================
from pyomo.environ import *
#from pyomo.opt import SolverFactory
#from pyomo.contrib.alternative_solutions.solnpool import gurobi_generate_solutions
from pyomo.opt import SolverManagerFactory

model = ConcreteModel()

# ---------------------------------Define Main and Additional Sets and Subsets ------------------------------
l_data = Sets_data.iloc[1, 2:4].values
g_data = Sets_data.iloc[2, 2:15].values 
p_data = Sets_data.iloc[3, 2:6].values
r_data = Sets_data.iloc[4, 2:6].values
s_data = Sets_data.iloc[5, 2:8].values
t_data = Sets_data.iloc[6, 2:8].values
d_data = Sets_data.iloc[7, 2:5].values
c_data = Sets_data.iloc[8, 2:8].values
h_data = Sets_data.iloc[9, 2:26].values
sc_data = Sets_data.iloc[10, 2:6].values
sv_data = Sets_data.iloc[11, 2:4].values
e_data = Sets_data.iloc[12, 2:5].values
I_data = Sets_data.iloc[1, 2:4].values
region1_data = Regions_data.iloc[2:48, 2].values
region2_data = Regions_data.iloc[2:48, 3].values
Neighbourhood_Regions = list(zip(region1_data,region2_data))

model.l = Set(initialize=['Trailer', 'Pipe'])
model.g = Set(initialize=g_data)
model.g1 = Set(initialize=g_data)
model.p = Set(initialize=p_data)
model.r = Set(initialize=r_data)
model.s = Set(initialize=s_data)
model.t = Set(initialize=[2,4,6])#t_data)
model.d1 = Set(initialize=d_data)
model.d2 = Set(initialize=[1, 2])
model.c = Set(initialize=c_data)
model.h = Set(initialize=h_data)
model.h1 = Set(initialize=h_data)
model.sc= Set(initialize=sc_data)
model.sv= Set(initialize=sv_data)
model.e = Set(initialize=e_data)
model.k=Set(initialize=["k" + str(i) for i in range(1, num_scenarios+1)])
model.o = Set(initialize=['Domestic', 'Commercial', 'Industrial', 'Transportation'])

Region3_data = Regions_data.iloc[6:32, 17].values
storage_data = Regions_data.iloc[6:32, 18].values


Region4_data = Regions_data.iloc[2:32, 17].values
storage1_data = Regions_data.iloc[2:32, 18].values

GS_data = list(zip(Region4_data, storage1_data))
GS_data1 = list(zip(Region3_data, storage_data))
GS_data2 = [('NO', 'OnTeeside'), ('NW', 'OnChesire'), ('NE', 'OnYorkshire'), ('NW', 'OffIrishSea')]

model.GS = Set(dimen=2, initialize=[(g,s) for g in model.g for s in model.s if (g,s) in GS_data])

model.GS1 = Set(dimen=2, initialize=[(g,sv) for g in model.g for sv in model.sv if (g,sv) in GS_data1])
model.GS2 = Set(dimen=2, initialize=[(g,sc) for g in model.g for sc in model.sc if (g,sc) in GS_data2])
Gimp_data = [(g_data[9]), (g_data[11]), (g_data[2]), (g_data[0])]
model.Gimp = Set(within=model.g, initialize= ['WS', 'SO', 'NO', 'NE', 'SC'])
#model.Gimp = Set(initialize=[(g) for g in model.g if (g) in Gimp_data])
GR_data=[(g_data[0], r_data[2]), (g_data[5], r_data[3]), (g_data[6], r_data[0])]
model.GR = Set(dimen=2, initialize=[(g,r) for g in model.g for r in model.r if (g,r) in GR_data])
model.N = Set(dimen=2, initialize=[(g,g1) for g in model.g for g1 in model.g if (g,g1) in Neighbourhood_Regions])


# Aliases
model.gg = Set(dimen=2, initialize=lambda model: [(g,g1) for g in model.g for g1 in model.g])
model.hh = Set(dimen=2, initialize=lambda model: [(h,h1) for h in model.h for h1 in model.h])


# ------ RangeSets -----

model.TT = Set(initialize=[2, 4, 6])
model.CC = RangeSet(1,5)  #  CC(c) /1*6/
model.HH = RangeSet(1, 24) #  HH(h) /1*24/


#%%
# -----------------Define Order in Pyomo for some variable ---------------
# %% Making order of set for some equations---------------

region_order = {region: i + 1 for i, region in enumerate(model.g)}
Trans_order = {transLine: i + 1 for i, transLine in enumerate(model.l)}  
Production_order = {production: i+1 for i, production in enumerate(model.p)}
Storage_order = {storage: i+1 for i, storage in enumerate(model.s)}
Cluster_order = {cluster: i+1 for i, cluster in enumerate(model.CC)}
Hour_order = {hour: i+1 for i, hour in enumerate(model.h)}

diameter_order1 = {diameter: i + 1 for i, diameter in enumerate(model.d1)}   
diameter_order2 = {diameter: i + 1 for i, diameter in enumerate(model.d2)}          
model.ord_d1 = Param(model.d1, initialize=diameter_order1)
model.ord_d2 = Param(model.d2, initialize=diameter_order2)

model.ord_g = Param(model.g, initialize=region_order)
#model.ord_d = Param(model.d, initialize=diameter_order)
model.ord_l = Param(model.l, initialize=Trans_order)
model.ord_p = Param(model.p, initialize=Production_order)
model.ord_s= Param(model.s, initialize=Storage_order )
model.ord_c= Param(model.CC,initialize=Cluster_order)
model.ord_h = Param(model.h,initialize=Hour_order)
# %% Assign spacific data for parameters-----------

DistSt_data = {(g, s): df_DistSt.iloc[i,2] 
          for i, g in enumerate(df_DistSt.iloc[:, 0])
          for j, s in enumerate(df_DistSt.iloc[:, 1])
          if i==j}

Data1 = Emissions_data.iloc[30:34, 3:7]
y_c_data = {(p,t): Data1.iloc[i,j]
            for i, p in enumerate(model.p)
            for j, t in enumerate(model.TT)}

Data2 = Emissions_data.iloc[21:25, 3:7]
y_e_data = {(p,t): Data2.iloc[i,j]
            for i, p in enumerate(model.p)
            for j, t in enumerate(model.t)}

diaH_data1 = H2Pipline_data.iloc[9:12, 1]
diaH_data = {(d1): diaH_data1.iloc[i]
             for i, d1 in enumerate(model.d1)}
'''
AV_data = {(c, h, g, e): df_Availability.iloc[i, 2 + 3 * g_idx + e_idx]
    for i, (c, h) in enumerate(zip(df_Availability.iloc[:, 0], df_Availability.iloc[:, 1]))  
    for g_idx, g in enumerate(model.g)  
    for e_idx, e in enumerate(model.e)}
'''


AV_data = {(c, h, e, g, k): df_Availability.iloc[i, 3 +  num_scenarios*g_idx]
    for i, (c, h,e) in enumerate(zip(df_Availability.iloc[:, 0], df_Availability.iloc[:, 1], df_Availability.iloc[:, 2]))  
    for g_idx, g in enumerate(model.g)  
    for k_idx, k in enumerate(model.k)}


                      
                               
df_Biomass.iloc[:, 0] = df_Biomass.iloc[:, 0].str.strip().str.upper()
br_data = dict(zip(df_Biomass.iloc[:, 0], df_Biomass.iloc[:, 1]))

cbio_data = dict(zip(df_bio.iloc[:, 0], df_bio.iloc[:, 1]))
cbio_data1 = {t: cost for t, cost in cbio_data.items() if t in {2, 4, 6}}
cccH_data = dict(zip(df_cccH.iloc[:, 0], df_cccH.iloc[:, 1]))
cccC_onshore_data = dict(zip(df_cccC_Onshore.iloc[:, 0], df_cccC_Onshore.iloc[:, 1]))
cccC_offshore_data = dict(zip(df_cccC_offshore.iloc[:, 0], df_cccC_offshore.iloc[:, 1]))

Data_price=gasprice_data.iloc[0:4,1:num_scenarios+1]
cgas_data = {(t,k): Data_price.iloc[i,j]
               for i, t in enumerate(model.TT)
               for j, k in enumerate(model.k)}

df_Cstart.iloc[:, 0] = df_Cstart.iloc[:, 0].str.strip().str.upper()
Cstart_data = dict(zip(df_Cstart.iloc[:, 0], df_Cstart.iloc[:, 1]))

df_Cshut.iloc[:, 0] = df_Cshut.iloc[:, 0].str.strip().str.upper()
Cshut_data = dict(zip(df_Cshut.iloc[:, 0], df_Cshut.iloc[:, 1]))

df_transposed= df_ct.T
df_transposed.columns = ['key', 'value']
ct_data = dict(zip(df_transposed['key'], df_transposed['value']))
ct_data1 = {t: cost for t, cost in ct_data.items() if t in {2, 4, 6}}

dc_data = dict(zip(df_dc.iloc[:,0], df_dc.iloc[:,1]))
dc_data1 = {t: cost for t, cost in dc_data.items() if t in {2, 4, 6}}
'''
DistPipe_data = {
    (g_row, g_col): df_DistPipe.iloc[i, j]
    for i, g_row in enumerate(model.g)
    for j, g_col in enumerate(model.g)}
'''
DistPipe_data = {
    (g_row, g_col): df_DistPipe.iloc[i, j]
    for i, g_row in enumerate(model.g)
    for j, g_col in enumerate(model.g)
    if df_DistPipe.iloc[i, j] > 0  
}

'''
Hsum_data={(h, h1, c): df_Hsum.iloc[i, 2 +c_idx] 
               for i, (h, h1) in enumerate(zip(df_Hsum.iloc[:, 0], df_Hsum.iloc[:, 1]))  
               for c_idx, c in enumerate(model.CC)  
               if df_Hsum.iloc[i, 2+c_idx] > 0}
'''
#Hsum_data={(h, c): df_Hsum.iloc[i, 2 +c_idx] 
               #for i, (h) in enumerate(zip( df_Hsum.iloc[:, 1]))  
               #for c_idx, c in enumerate(model.CC)  
               #if df_Hsum.iloc[i, 2+c_idx] > 0}

DistRes_data = {(g, r): df_DistRes.iloc[i,2] 
          for i, g in enumerate(df_DistRes.iloc[:, 0])
          for j, r in enumerate(df_DistRes.iloc[:, 1])
          if i==j}


Dist_data = {
    (g_row, g_col): df_Dist.iloc[i, j]
    for i, g_row in enumerate(model.g)
    for j, g_col in enumerate(model.g) 
    if df_Dist.iloc[i, j] > 0}
    

DT_data = dict(zip(df_DT.iloc[:, 0], df_DT.iloc[:, 1]))
df_ec_transposed= df_ec.T
df_ec_transposed.columns = ['key', 'value']
ec_data = dict(zip(df_ec_transposed['key'], df_ec_transposed['value']))
ec_data1 = {t: cost for t, cost in ec_data.items() if t in {2, 4, 6}}
Data4 = Production_data.iloc[69:73, 3:7]
eta_data = {(p,t): Data4.iloc[i,j]
            for i, p in enumerate(model.p) if model.ord_p[p]<4
            for j, t in enumerate(model.t)}

eta_dataWE=etaWE_data.iloc[0:4,1:num_scenarios+1]
eta_dataWE = {(t,k): eta_dataWE.iloc[i,j]
            for i, t in enumerate(model.TT) 
            for j, k in enumerate(model.k)}


df_emtarget_transposed= df_emtarget.T
df_emtarget_transposed.columns = ['key', 'value']
emtarget_data = dict(zip(df_emtarget_transposed['key'], df_emtarget_transposed['value']))
emtarget_data1 = {t: emission for t, emission in emtarget_data.items() if t in {2, 4, 6}}
#GasDem_data = {(c, h, g): df_GasDem.iloc[i, 2+j] 
               #for i, (c, h) in enumerate(zip(df_GasDem.iloc[:, 0], df_GasDem.iloc[:, 1]))  
               #for j, g in enumerate(model.g)}

Dom_data={(c, h, g,k): df_DomDem.iloc[i, 2 +  num_scenarios*g_idx] 
               for i, (c, h) in enumerate(zip(df_DomDem.iloc[:, 0], df_DomDem.iloc[:, 1]))  
               for g_idx, g in enumerate(model.g)  
               for k_idx, k in enumerate(model.k)}

Com_data={(c, h, g,k): df_ComDem.iloc[i, 2 +  num_scenarios*g_idx] 
               for i, (c, h) in enumerate(zip(df_ComDem.iloc[:, 0], df_ComDem.iloc[:, 1]))  
               for g_idx, g in enumerate(model.g)  
               for k_idx, k in enumerate(model.k)}
Ind_data={(c, h, g,k): df_IndDem.iloc[i, 2 +  num_scenarios*g_idx] 
               for i, (c, h) in enumerate(zip(df_IndDem.iloc[:, 0], df_IndDem.iloc[:, 1]))  
               for g_idx, g in enumerate(model.g)  
               for k_idx, k in enumerate(model.k)}

hdem_data = {
    (o, t, k): df_hdem.iloc[i, 2 + k_idx]
    for i, (o, t) in enumerate(zip(df_hdem.iloc[:, 0], df_hdem.iloc[:, 1]))
    for k_idx, k in enumerate(model.k)
}



Data5 = Renewables_data.iloc[28:31, 1:14]
landAV_data = {(e,g): Data5.iloc[i,j]
            for i, e in enumerate(model.e)
            for j, g in enumerate(model.g)}

Data6 = Production_data.iloc[34:38, 5]
Capmax_data = {(p): Data6.iloc[i]
            for i, p in enumerate(model.p)}

Data7 = Production_data.iloc[34:38, 1]
Capmin_data = {(p): Data7.iloc[i]
            for i, p in enumerate(model.p)}

Data8 = Production_data.iloc[4:8, 2:7]

selected_t_indices = [0, 2, 4]  # Corresponding to t = 2, 4, 6 (Python uses zero-based indexing)
Data8_filtered = Data8.iloc[:, selected_t_indices]  

# Corrected dictionary comprehension with filtering for t ∈ {2,4,6}
pccost_data = {(p, t): Data8_filtered.iloc[i, j]
               for i, p in enumerate(model.p) if model.ord_p[p] < 4
               for j, t in enumerate(model.t) if t in {2, 4, 6}}


DatapccostWE=pccostWE_data.iloc[0:4,1:num_scenarios+1]
pccost_data_WE = {(t,k): DatapccostWE.iloc[i,j]
            for i, t in enumerate(model.TT)
            for j, k in enumerate(model.k)}

Data9 = Production_data.iloc[14:18, 2:7]
Data9_filtered = Data9.iloc[:, selected_t_indices] 
pocostF_data = {(p, t): Data9_filtered.iloc[i, j]
               for i, p in enumerate(model.p) if model.ord_p[p] < 4
               for j, t in enumerate(model.t) if t in {2, 4, 6}}

DatapocostWE = pocostFWE_data.iloc[0:4, 1:num_scenarios+1]
pocostF_dataWE = {(t,k): DatapocostWE.iloc[i,j]
            for i, t in enumerate(model.TT) 
            for j, k in enumerate(model.k)}

Data10 = Production_data.iloc[14:18, 11:16]
Data10_filtered = Data10.iloc[:, selected_t_indices] 
pocostV_data = {(p, t): Data10_filtered.iloc[i, j]
               for i, p in enumerate(model.p) if model.ord_p[p] < 4
               for j, t in enumerate(model.t) if t in {2, 4, 6}}

DatapocostVWE = pccosVWE_data.iloc[0:4, 1:num_scenarios+1]
pocostV_dataWE = {(t,k): DatapocostVWE.iloc[i,j]
            for i, t in enumerate(model.TT) 
            for j, k in enumerate(model.k)}


Data11 = H2Pipline_data.iloc[20:23, 1]
qHmax_data = {(d1): Data11.iloc[i]
            for i, d1 in enumerate(model.d1)}

Data12 = CO2Pipline_data.iloc[49:51, 1]
qCmax_data = {(d2): Data12.iloc[i]
            for i, d2 in enumerate(model.d2)}

Data13 = Storage_data.iloc[59:65, 1]
QImax_data = {(s): Data13.iloc[i]
              for i, s in enumerate(model.s)}

Data14 = Storage_data.iloc[59:65, 4]
QRmax_data = {(s): Data14.iloc[i]
              for i, s in enumerate(model.s)}

Data15 = CO2Reservior_data.iloc[3:7, 4]
rcap_data = {(r):Data15.iloc[i]
             for i, r in enumerate(model.r)}

Data16 = CO2Reservior_data.iloc[12:16, 1]
ri0_data = {(r):Data16.iloc[i]
             for i, r in enumerate(model.r)}

Data17 = Production_data.iloc[50:54, 8]
RD_data = {(p): Data17.iloc[i]
           for i, p in enumerate(model.p)}

Data18 = Storage_data.iloc[36:42, 5]
scap_max_data = {(s): Data18.iloc[i]
                 for i, s in enumerate(model.s)}


rccost_data = {(e, t, k): df_rccost.iloc[i, 2+k_idx]
    for i, (e, t) in enumerate(zip(df_rccost.iloc[:, 0], df_rccost.iloc[:, 1]))    
    for k_idx, k in enumerate(model.k)}

Data_prob = Prob_data.iloc[0:num_scenarios+1,1]
prob_data={(k): Data_prob.iloc[i]
               for i, k in enumerate(model.k)}

#rccost = Renewables_data.iloc[3:6, 3:7]
#rccost_data = {(e,t): rccost.iloc[i,j]
              # for i, e in enumerate(model.e) 
              # for j, t in enumerate(model.t)}

rocost = Renewables_data.iloc[10:13, 2:7]
rocost_data = {(e,t): rocost.iloc[i,j]
               for i, e in enumerate(model.e)
               for j, t in enumerate(model.t) if t in {2, 4, 6}}

Data19 = Storage_data.iloc[2:8, 1] 
sccost_data = {(s): Data19.iloc[i]
               for i, s in enumerate(model.s)}

Data20 = Storage_data.iloc[14:20, 1]
socostF_data = {(s): Data20.iloc[i]
                for i, s in enumerate(model.s)}

Data21 = Storage_data.iloc[14:20, 7]
socostV_data = {(s): Data21.iloc[i]
                for i, s in enumerate(model.s)}

Data22 = Production_data.iloc[23:27, 1]
Pcap_data = {(p): Data22.iloc[i]
             for i, p in enumerate(model.p)}

Data23 = Storage_data.iloc[25:31, 1]
SCap_data = {(s): Data23.iloc[i]
             for i, s in enumerate(model.s)}

Data24 = Production_data.iloc[50:54, 1]
UT_data = {(p): Data24.iloc[i]
           for i, p in enumerate(model.p)}


Data25 = Biomass_data.iloc[0:4, 1:num_scenarios+1]
Vbio_data = {(t,k): Data25.iloc[i,j]
            for i, t in enumerate(model.TT)
            for j, k in enumerate(model.k)}


#Data26 = Demand_data.iloc[1:7, 17]
#WF_data = {(c): Data26.iloc[i]
          # for i, c in enumerate(model.CC)}

Cluster_data1=Cluster_data.iloc[3:8,1]
WF_data1 = {(c): Cluster_data1.iloc[i]
           for i, c in enumerate(model.CC)}

# %%
#======================================Parameters========================================================
# %% Define Parameter------------------------------
model.beta = Param(initialize=0.15, doc='Ratio of stored amount (%)')

# Distance between region and underground storage
model.DistSt = Param(model.g, model.sc, initialize=DistSt_data, doc='distance between region g and underground storage type s')

# CO2 capture and emission coefficients
model.y_c = Param(model.p, model.t, initialize=y_c_data, doc='CO2 capture coefficient for plant type p in time period t (tn CO2 / MWh H2)')
model.y_e = Param(model.p, model.t, initialize=y_e_data,   doc='CO2 emission coefficient for plant type p and size j in time period t (tn CO2 / MWh H2)')


# Pipeline operating cost ratios
model.deltaH = Param(initialize=0.05, doc='Ratio of hydrogen regional pipeline operating costs to capital costs (%)')
model.deltaC_onshore = Param(initialize=0.05, doc='Ratio of onshore CO2 pipeline operating costs to capital costs')
model.deltaC_offshore = Param(initialize=0.05, doc='Ratio of offshore CO2 pipeline operating costs to capital costs')


# Pipeline diameters
model.diaH = Param(model.d1, initialize=diaH_data)
model.diaC_onshore = Param(model.d2, initialize={1: 0.6, 2: 1.2}, doc='Diameter of an onshore CO2 pipeline of diameter size d (m)')
model.diaC_offshore = Param(model.d2, initialize={1: 0.6, 2: 1.2}, doc='Diameter of an offshore CO2 pipeline of diameter size d (m)')

# Hydrogen import ratio
model.iota = Param(initialize=0.1, doc='Maximum percentage of international hydrogen imports over the total demand (%)')

# Time-related parameters
model.dur = Param(initialize=10, doc='Duration of time periods (y)')
model.LTonshore = Param(initialize=50, doc='Useful life of onshore CO2 pipelines (y)')
model.LToffshore = Param(initialize=50, doc='Useful life of offshore CO2 pipelines (y)')
model.LTpipe = Param(initialize=50, doc='Useful life of hydrogen pipelines (y)')
model.a = Param(initialize=365, doc='Days in a year (days)')

model.LTp = Param(model.p, initialize={'SMRCCS':40, 'ATRCCS':40, 'BECCS':30, 'WE':30},doc='Useful life of hydrogen production plants (y)')
model.LTs = Param(model.s, initialize={'OnTeeside':40, 'OnChesire':40, 'OnYorkshire':40, 'OffIrishSea':40, 'MPSV':40, 'HPSV':40}, doc='Useful life of hydrogen storage facilities (y)')
model.LTt = Param(model.l, initialize={'Trailer': 15}, doc='Useful life of hydrogen road transportation modes (y)')



# Biomass parameters
model.br = Param(model.g, initialize=br_data,doc='Parameter for region-specific values')
model.bp = Param( initialize=0.5)
model.cbio = Param(model.t, initialize=cbio_data1, doc='Biomass cost in time period t (€/MWh)')


# Pipeline costs and renewable energy parameters
model.cccH = Param(model.d1, initialize=cccH_data, doc='Capital costs of a regional hydrogen pipeline of diameter size q d (€/k km-1)')
model.cccC_onshore = Param(model.d2, initialize=cccC_onshore_data, doc='Capital costs of an onshore CO2 pipeline of diameter size d (€/k km-1)')
model.cccC_offshore = Param(model.d2, initialize=cccC_offshore_data, doc='Capital costs of an offshore CO2 pipeline of diameter size d (€/k km-1)')
model.cgas = Param(model.TT,model.k, initialize=cgas_data, doc='Natural gas cost in time period t (€/MWh)')
model.crf = Param(initialize=0.07, doc='Capital recovery factor')

# Start-up and shut-down costs for technologies
model.Cstart = Param(model.p, initialize=Cstart_data, doc='Cost for starting up for each technology type (€/MW)')
model.Cshut = Param(model.p,initialize=Cshut_data, doc='Cost for shutting down for each technology type (€/MW)')

# Carbon tax and demand parameters
model.ct = Param(model.t, initialize=ct_data1,doc='carbon tax i time period t (€/kg CO2)')
model.dc = Param(model.t, initialize=dc_data1, doc='Demand coefficient at time period t')
#model.dem = Param(model.g, model.t, model.c, model.h, doc='Total hydrogen demand in region g in time period t (MW)')
model.hdem=Param(model.o,model.TT,model.k, initialize=hdem_data)

model.Com=Param(model.CC,model.h,model.g,model.k, initialize=Com_data)
model.Dom=Param(model.CC,model.h,model.g,model.k, initialize=Dom_data)
model.Ind=Param(model.CC,model.h,model.g,model.k, initialize=Ind_data)


# Transportation and pipeline parameters
model.dw = Param(model.l, initialize={'Trailer':16.62 }, doc='Driver wage of road transportation mode l (€/h)')
model.DistPipe = Param(model.g, model.g, initialize=DistPipe_data, within=NonNegativeReals, doc='Delivery distance of an onshore CO2 pipeline between regions g and g1 (km)')
model.DistRes = Param(model.g, model.r, initialize=DistRes_data, doc='Distance from CO2 collection point in region g to reservoir r (km)')
model.Dist = Param(model.g, model.g, initialize=Dist_data, doc='Regional delivery distance of hydrogen transportation mode l in region g (km)')


# Technical parameters for plants and pipelines
model.DT = Param(model.p, initialize=DT_data, doc='Min down time (h)')
model.ec = Param(model.t, initialize=ec_data1, doc='Cost of electricity back to grid (€/MWe)')
model.eta = Param(model.p, model.TT, initialize=eta_data, doc='Efficiency of WE in time period t (%)')
model.etaWE=Param(model.TT,model.k, initialize=eta_dataWE)
model.emtarget = Param(model.t, initialize=emtarget_data1, doc='Emissions target in time period t (kgCO2)')

# Road transportation costs and fuel economy
model.feR = Param(model.l, initialize={'Trailer': 2.3}, doc='Fuel economy of road transportation mode l transporting product type i within a region (km/l)')
model.fp = Param(model.l, initialize={'Trailer': 1.63 }, doc='Fuel price of road transportation mode l (€/l)')
#model.GasDem = Param(model.CC, model.h, model.g,  initialize=GasDem_data, doc='Hydrogen demand for each region g each cluster c and hour h (MWh)')
model.ge = Param(model.l, initialize={'Trailer': 0.25 }, doc='General expenses of road transportation mode l transporting product type i (€/d)')

# Economic parameters
model.ir = Param(initialize=0.06, doc='Discount rate (%)')
model.landAV = Param(model.e, model.g, initialize=landAV_data, doc='Land availability of renewable e in region g (MW)')
model.lut = Param(model.l, initialize={'Trailer':2}, doc='Load and unload time of road transportation mode l (h)')
model.me = Param(model.l, initialize={'Trailer':0.07}, doc='Maintenance expenses of road transportation mode l (€/km)')
model.nel = Param(initialize=30, doc='Economic life cycle of capital investments (y)')

# Initial number of plants and storage units
model.np0 = Param(model.p, model.g, initialize=0, doc='Initial number of hydrogen production plants of technology p and size j in region g')
model.ns0 = Param(model.s, model.g, initialize=0, doc='Initial number of hydrogen storage facilities of type s and size j in region g')

# Production and storage capacity parameters
model.pcap_max = Param(model.p, initialize=Capmax_data, doc='Maximum capacity of a hydrogen production plant of type p and size j (MW)')
model.pcap_min = Param(model.p, initialize=Capmin_data, doc='Minimum capacity of a hydrogen production plant of type p and size j (MW)')
model.pccost = Param(model.p, model.t, initialize=pccost_data, doc='Capital cost of a production plant of type p (€/kW)')
model.pccostWE = Param(model.TT, model.k, initialize=pccost_data_WE, doc='Capital cost of a production plant of type p (€/kW)')

model.pimp = Param(initialize=127.6, doc='Price of hydrogen import (€/MWh)')
model.pocostF = Param(model.p, model.t, initialize=pocostF_data, doc='Operating production cost in a production plant of type p (€/MWh/y)')
model.pocostFWE=Param(model.TT, model.k, initialize=pocostF_dataWE)
model.pocostV = Param(model.p, model.t, initialize=pocostV_data, doc='Operating production cost in a production plant of type p (€/MWh)')
model.pocostVWE = Param( model.TT, model.k, initialize=pocostV_dataWE, doc='Operating production cost in a production plant of type p (€/MWh)')

# Flow rate and capacity limits
model.qHmax = Param(model.d1, initialize=qHmax_data, doc='Maximum flow rate in a hydrogen pipeline of diameter size d (kg H2/day)')
model.qCmax = Param(model.d2, initialize=qCmax_data, within=NonNegativeReals, doc='Maximum flow rate in a CO2 pipeline of diameter size d (kg H2/day)')
model.QImax = Param(model.s, initialize=QImax_data, doc='Maximum injection rate for each storage type s')
model.QRmax = Param(model.s, initialize=QRmax_data, doc='Maximum retrieval rate for each storage type s')

# Reservoir-related parameters
model.rcap = Param(model.r, initialize=rcap_data, doc='Total capacity of reservoir r (kg CO2-eq)')
model.ri0 = Param(model.r, initialize=ri0_data, doc='Initial CO2 inventory in reservoir r (kg CO2)')

# Ramp-up and ramp-down parameters
model.RD = Param(model.p, initialize=RD_data, doc='Commit Ramp down')
model.rccost = Param(model.e, model.TT,model.k, initialize=rccost_data, doc='Renewable e capital cost in time period t (€/MW)')
model.rocost = Param(model.e, model.t, initialize=rocost_data, doc='Renewable e operating cost in time period t (€/MW)')

# Storage parameters
model.RU = Param(model.p, initialize=RD_data, doc='Commit Ramp up', )
model.scap_max = Param(model.s, initialize=scap_max_data, doc='Maximum capacity of a storage facility of type s (MWh H2)')
model.scap_min = Param(model.s, initialize=0, doc='Minimum capacity of a storage facility of type s (MWh H2)')
model.sccost = Param(model.s, initialize=sccost_data, doc='Fixed operating storage cost in a production plant of type p (€/MW/y)')
model.socostF = Param(model.s, initialize=socostF_data, doc='Fixed operating storage cost in a production plant of type p (€/MW/y)')
model.socostV = Param(model.s, initialize=socostV_data, doc='Variable operating storage cost in a production plant of type p (€/kWh stored)')

# Road transportation speed and capacity
model.spR = Param(model.l, initialize={'Trailer': 55}, doc='Regional average speed of road transportation mode l (km/h)')
model.st0 = Param(model.s, model.g, initialize=0, doc='Storage at time 0')
model.tcap = Param(model.l, initialize={'Trailer': 21.66 }, doc='Capacity of road transportation mode l transporting product type i (MWh unit-1)')
model.tmc = Param(model.l, initialize={'Trailer': 253000 }, doc='Capital cost of establishing a road transportation unit of transportation mode l (€/unit)')
model.tmaR = Param(model.l, initialize={'Trailer': 18 }, doc='Regional availability of road transportation mode l (h/day)')

# Unit capacity for production and storage
model.PCap = Param(model.p, initialize=Pcap_data, doc='Unit capacity for production type p (MW)')
model.SCap = Param(model.s, initialize=SCap_data, doc='Unit capacity for storage type s (MW)')

# Initial operating units
model.uInit = Param(model.p, model.g, model.t, initialize=0, doc='Initial operating units type p in region g at time period t')

# Technical parameters for up and down time
model.UT = Param(model.p, initialize=UT_data, doc='Min up time (h)')

# Biomass consumption and cluster weights
model.Vbio_max = Param(model.TT,model.k, initialize=Vbio_data, doc='Maximum biomass consumption in year t')
model.WF = Param(model.CC, initialize=WF_data1, doc='Weight of clusters')
#model.WF = Param(model.CC, initialize={k: v * 0.2526 for k, v in list(WF_data.items())[1:]}, doc='Weight of clusters')

# ---- Scalar ----
model.y1 = Param(initialize=2, doc="Scalar y1")
model.hf=Param(initialize=n_slice)



# Define a function for initializing the values of dfc
def dfc_init(model, t):
    return round(1 / (1 + model.ir) ** (model.dur * t - model.dur), 2)

model.dfc = Param(model.TT, initialize=dfc_init,  doc='Discount factor for capital costs in time period t')

#model.dfc = Expression(model.t, rule=lambda model, t: round(1 / (1 + model.ir) ** (model.dur * t - model.dur), 2), doc='Discount factor for capital costs in time period t')

def dfo_init(model,t):
    return round(
        1 / (1 + model.ir) ** (model.dur * t - 5) +
        1 / (1 + model.ir) ** (5 * t - 4) +
        1 / (1 + model.ir) ** (5 * t - 3) +
        1 / (1 + model.ir) ** (5 * t - 2) +
        1 / (1 + model.ir) ** (5 * t - 1),
        2
    )
model.dfo= Param(model.TT, initialize=dfo_init)

def biomass_availability_init(model, g, t,k):
    return model.bp * model.br[g] * model.Vbio_max[t,k] * 1000000

model.BA = Param(model.g, model.TT, model.k, initialize=biomass_availability_init,  doc="Biomass availability in region g and time period t")


'''
model.BA = Expression(model.g, model.t, rule=lambda model, g, t: model.bp * model.br[g] * model.Vbio_max[t] * 1000000,  doc='Biomass availability in region g and time period t')
'''
#model.dem = Expression(model.g, model.t, model.CC, model.h, rule=lambda model, g, t, c, h:model.dc[t]*model.GasDem[c,h,g])
CH_data = {
    1: list(range(1, 25)),
    2: list(range(1, n_slice+1)),
    3: list(range(1, n_slice+1)),
    4: list(range(1, n_slice+1)),
    5: list(range(1, n_slice+1))
}

# Define CH as a set indexed by C
model.CH = Set(dimen=2, initialize=[(c,h) for c in CH_data for h in CH_data[c]])

def build_dem_data(model):
    dem_data = {}
    for g in model.g:
        for t in model.TT:
            for c in model.c:
                for h in model.h:
                    if (c, h) in model.CH:
                        for k in model.k:
                            value = (
                                model.hdem['Domestic', t, k] * model.Dom[c, h, g, k] / 288.6 +
                                model.hdem['Commercial', t, k] * model.Com[c, h, g, k] / 94.5 +
                                model.hdem['Industrial', t, k] * model.Ind[c, h, g, k] / 89.4 +
                                (model.hdem['Transportation', t, k] * 1000000 / 8760 / len(model.g))
                            )
                            dem_data[g, t, c, h, k] = value
    return dem_data

model.dem = Param(model.g, model.TT, model.c, model.h, model.k, initialize=build_dem_data)





'''
model.RI_up = Expression(model.r, model.TT, rule=lambda model, r,t :model.rcap[r]/1000 )
'''
'''
model.Npipe = Set(dimen=2, initialize=[(g,g1) for g in model.g for g1 in model.g if model.DistPipe[g, g1] > 0 ])
'''
model.Npipe = Set(dimen=2,initialize=model.DistPipe.keys(), doc='Set of region pairs with nonzero pipeline distances')

'''
model.Npipe = Set(
    initialize=lambda model: [(g, g1) for g in model.g for g1 in model.g1 if model.DistPipe[g, g1] > 0],
    doc="Set of region pairs (g, g1) with a non-zero distance (DistPipe > 0)"
)
'''

# Availability and initial availability parameters
model.AV = Param(model.CC, model.h, model.e, model.g,model.k, initialize=AV_data, doc='Availability of renewable e in region g, cluster c and hour h (%)')
model.ayHR0 = Param(model.d1, model.Npipe,initialize=0, doc='Initial availability of a regional hydrogen pipeline of diameter size d between regions g and g1 (0-1)')
model.ayC0 = Param(model.d2, model.N, initialize=0, doc='Initial availability of an onshore CO2 pipeline of diameter size d between regions g and g1 (0-1)')
model.aeC0 = Param(model.r, initialize=0, doc='Initial availability of an offshore CO2 pipeline between collection point in regions g and reservoir r (0-1)')

pccostK_data = {}

for p in model.p:
    for t in model.TT:
        for k in model.k:
            if model.ord_p[p] < 4:
                pccostK_data[(p, t, k)] = model.pccost[p, t]
            elif model.ord_p[p] == 4:
                pccostK_data[(p, t, k)] = model.pccostWE[t, k] / 1000

model.pccostK=Param(model.p,model.t,model.k, initialize=pccostK_data)

pocostKF_data={}
for p in model.p:
    for t in model.TT:
        for k in model.k:
            if model.ord_p[p] < 4:
                pocostKF_data[(p, t, k)] = model.pocostF[p, t]
            elif model.ord_p[p] == 4:
                pocostKF_data[(p, t, k)] = model.pocostFWE[t, k]

model.pocostKF=Param(model.p,model.t,model.k, initialize=pocostKF_data)

pocostKV_data={}
for p in model.p:
    for t in model.TT:
        for k in model.k:
            if model.ord_p[p] < 4:
                pocostKV_data[(p, t, k)] = model.pocostV[p, t]
            elif model.ord_p[p] == 4:
                pocostKV_data[(p, t, k)] = model.pocostVWE[t, k]

model.pocostKV=Param(model.p,model.t,model.k, initialize=pocostKV_data)

etaK_data={}
for p in model.p:
    for t in model.TT:
        for k in model.k:
            if model.ord_p[p] < 4:
                etaK_data[(p, t, k)] = model.eta[p, t]
            elif model.ord_p[p] == 4:
                etaK_data[(p, t, k)] = model.etaWE[t, k]

model.etaK=Param(model.p,model.t,model.k, initialize=etaK_data)

model.prob=Param(model.k, initialize=prob_data)



theta_data={(c, h): df_theta.iloc[i, 2]
    for i, (c, h) in enumerate(zip(df_theta.iloc[:, 0], df_theta.iloc[:, 1])) }
         
model.theta = Param(model.CC,model.h, initialize=theta_data, doc="Scalar theta")




hsum_index = list(df_Hsum.itertuples(index=False, name=None))


model.Hsum = Set(dimen=3, initialize=hsum_index)

'''
InvP_data = {}
for g in model.g:
    for t in model.TT:
        InvP_data[('SMRCCS', g, t)] = 10
        InvP_data[('ATRCCS', g, t)] = 10
        InvP_data[('BECCS', g, t)] = 10
        InvP_data[('WE', g, t)] = 50
model.InvP_up = Param(model.p, model.g, model.TT, initialize=InvP_data)
'''

InvP_bounds = {
    'SMRCCS': 10,
    'ATRCCS': 10,
    'BECCS': 10,
    'WE': 50
}

def InvP_bounds_rule(model, p, g, t):
    if p in InvP_bounds and t in model.TT:
        return (0, InvP_bounds[p]) 
    

InvS_bounds = {
    'MPSV': 80,
    'HPSV': 80}
def InvS_bounds_rule(model, s, g, t):
    if s in InvS_bounds and t in model.TT:
        return (0, InvS_bounds[s])
    elif s in ['OnTeeside', 'OnChesire', 'OnYorkshire', 'OffIrishSea']:
        return (0, 1) 
    return (0, None)  

def NS_bounds_rule(model, s, g, t):
    if s in model.sc and (g, s) in model.GS2 and t in model.TT:
        return (0, 1)  
    else:
        return (0, None)
    
def RI_bounds_rule(model, r, t,k):
    return (0, model.rcap[r] / 1000)


def Qup_bounds_rule(model, l, g, g1,t, c, h,k):
    if l in ['Pipe'] and (c,h) in model.CH and (g,g1) in model.Npipe:
        return (0,15343)
    return (0, None)


def demH_init(model, g, t, c, HH, k):
    return sum(
        (model.dem[g, t, c, h, k] / model.theta[c, h]) if model.theta[c, h] != 0 else 0
        for h in model.HH if ((HH, h, c) in model.Hsum)
    )

model.demH = Param(model.g, model.TT, model.CC, model.HH, model.k, initialize=demH_init)

 
def AVH_init(model, c, HH, g, e, k):
    return sum(
        (model.AV[c, h, e, g, k] / model.theta[c, h]) if model.theta[c, h] != 0 else 0
        for h in model.HH if ((HH,h, c) in model.Hsum)
    )

model.AVH = Param(model.CC, model.HH, model.g, model.e, model.k, initialize=AVH_init)
  





# %%
#======================================Variables========================================================
# %% Define Variables------------------------------

model.InvP = Var(model.p, model.g, model.TT, within=NonNegativeIntegers, bounds=InvP_bounds_rule,
                 doc="Investment of new plants of type p producing in region g in time period t")
#.InvP_up = Var(model.p, model.g, model.TT, within=NonNegativeIntegers)
model.InvS = Var(model.s, model.g, model.TT, within=NonNegativeIntegers, bounds=InvS_bounds_rule,
                 doc="Investment of new storage facilities of type in region g in time period t")
#model.InvS_up = Var(model.s, model.g, model.TT,  within=NonNegativeIntegers)
model.Yh = Var(model.d1, model.Npipe, model.TT, within=NonNegativeIntegers,  doc="Establishment of hydrogen pipelines of diameter size d for regional distribution in region g in time period t")
model.Yon = Var(model.d2, model.N, model.TT, within=NonNegativeIntegers, 
                doc="Establishment of onshore CO2 pipelines of diameter size d in region g in time period t")
model.Yoff = Var(model.d2, model.GR, model.TT, within=NonNegativeIntegers, 
                 doc="Establishment of offshore CO2 pipelines of diameter size d in region g in time period t")
model.Yst = Var(model.d1, model.GS2, model.TT, within=NonNegativeIntegers, 
                doc="Establishment of hydrogen pipelines of diameter size d in region g in time period d to storage type s")


# Positive variables
model.NP = Var(model.p, model.g, model.TT, within=NonNegativeReals, doc="Number of plants of type j and size p in region g in time period t")
model.NS = Var(model.s, model.g, model.TT, within=NonNegativeReals, bounds=NS_bounds_rule, doc="Number of storage facilities of type s and size p in region g in time period t")
#model.NS_up = Var(model.s, model.g, model.TT, within=NonNegativeIntegers)
#model.ITU = Var(['Trailer'], model.Npipe, model.TT, within=NonNegativeReals, bounds=(0,25), doc="Number of new transportation units of type l for regional transportation byroad in region g to region g acquired in time period t")
#model.NTU = Var(['Trailer'], model.Npipe, model.TT, within=NonNegativeReals,   doc="Number of transportation units of type l for regional transportation by road in region g in time period t")
model.AY = Var(model.d1, model.Npipe, model.TT, within=NonNegativeReals,  doc="availability of hydrogen pipelines of diameter size d for regional distribution in region g in time period t")
model.AYon=Var(model.d2, model.N, model.TT, within=NonNegativeReals, doc="availability of onshore CO2 pipelines of diameter size d for local distribution in region g in time period t")
model.AYoff = Var(model.d2, model.GR, model.TT, within=NonNegativeReals, doc="availability of offshore CO2 pipelines of diameter size d for local distribution in region g in time period t")
model.AYst = Var(model.d1, model.GS2, model.TT, within=NonNegativeReals, doc="availability of hydrogen pipelines of diameter size d for distribution in region g in time period t")
model.CL = Var(model.g, model.TT, model.CC, model.HH, model.k, within=NonNegativeReals, doc='Curtailment (MW)')
model.InvR = Var(model.e, model.g, model.TT, within=NonNegativeReals, bounds=(0, 10000), doc='Invested capacity of renewable (MW)')
#model.InvR_up = Var(model.e, model.g, model.TT, within=NonNegativeReals, bounds=(0, 10000))
model.IMP = Var(model.g, model.TT, model.CC, model.h,model.k, within=NonNegativeReals,
                doc='Flow rate of international import (MW)')
model.NR = Var(model.e, model.g, model.TT, within=NonNegativeReals,  doc='Capacity of renewable (MW)')
model.Pr = Var(model.p, model.g, model.TT, model.CC,model.HH, model.k, within=NonNegativeReals,      doc='Production rate (MW)')
model.Pre = Var(model.e, model.g, model.TT, model.CC, model.HH, model.k, within=NonNegativeReals,    doc='Electricity production from renewable (MW)')
model.Q = Var(model.l, model.Npipe, model.TT, model.CC, model.HH,model.k,  within=NonNegativeReals, bounds=Qup_bounds_rule, doc='Regional flowrate of H2 (MWh)')
model.Qi = Var(model.g,model.s, model.TT, model.CC,model.HH,model.k,  within=NonNegativeReals,  doc='H2 via pipeline to storage (MWh)')
model.Qr = Var(model.s, model.g, model.TT, model.CC, model.HH,model.k,  within=NonNegativeReals,  doc= 'flowrate of H2 via pipeline from region g to storage type s in time period t(MWh)')
model.Qon = Var(model.N, model.TT, model.CC, model.HH, model.k, within=NonNegativeReals, bounds=(0,1.17E+04), doc='Flowrate of CO2 via onshore pipelines (kg CO2/d)')
model.Qoff = Var(model.GR, model.TT, model.CC, model.HH, model.k, within=NonNegativeReals, bounds=(0,1.17E+04), doc='Flowrate of CO2 via offshore pipelines (kg CO2/d)')
model.BS=Var(model.s,model.g,model.t,model.CC,model.k)
#odel.Rdown = Var(model.p, model.g, model.TT, model.CC, model.h, within=NonNegativeReals,initialize=0, doc='Upward reserve contribution (MWh)')
#odel.Rup = Var(model.p, model.g, model.TT, model.CC, model.h, within=NonNegativeReals, doc='Downward reserve contribution (MWh)')
model.St = Var(model.s,model.g, model.TT, model.CC, model.HH, model.k, within=NonNegativeReals, doc='Average inventory of product stored (kW)')
model.Vbio = Var(model.TT, model.k,within=NonNegativeReals,  doc='Biomass consumption (kg)' )
model.Vgas = Var(model.TT, model.k, within=NonNegativeReals,  doc='Gas consumption (kg)')
model.slak1 = Var(model.g, model.TT, model.CC, model.h, model.k, within=NonNegativeReals, doc='Slack variable 1')
#model.slak2 = Var(model.g, model.TT, model.CC, model.h, within=NonNegativeReals, doc='Slack variable 2')
model.PCC = Var(model.k,within=NonNegativeReals)
model.SCC = Var()
model.TCC=Var()
model.POC = Var(model.k,within=NonNegativeReals)
model.SOC = Var(model.k,within=NonNegativeReals)
model.RI = Var(model.r, model.TT, model.k, within=NonNegativeReals, bounds=RI_bounds_rule)
#model.RI_up = Var(model.r, model.TT, within=NonNegativeReals)
#model.TC = Var()
#model.RCC = Var()
#model.FCR = Var()
#model.GCR = Var()
#model.LCR = Var()
#model.MCR = Var()
model.PipeOC = Var()
model.PipeCC = Var()
model.CEC = Var(model.k)
model.IIC = Var(model.k,within=NonNegativeReals)
model.ReC = Var(model.k,within=NonNegativeReals)
model.GC = Var(model.k,within=NonNegativeReals )
model.BC = Var( model.k,within=NonNegativeReals)
model.ROC = Var(model.k,within=NonNegativeReals)
model.TOC=Var(model.k, within=NonNegativeReals)
model.em = Var(model.TT,model.k, within=Reals)
model.TC = Var(model.k)



# %%
#======================================Objective Function========================================================*
# %% Objective Function and related components----

# Constraint for PCC
def pcc_rule(model,k):
    return 0.001*model.PCC[k] == 0.001*sum(
        model.dfc[t] *model.pccostK[p, t,k] * model.PCap[p] * model.InvP[p, g, t]
        for p in model.p for g in model.g for t in model.TT
    )
model.PCCConstraint = Constraint(model.k, rule=pcc_rule)

# Constraint for SCC
def scc_rule(model):
    return 0.001*model.SCC ==  0.001*sum(
        model.dfc[t] * model.sccost[s] * model.SCap[s] * model.InvS[s, g, t]
       for s in model.s  for g in model.g if  (g, s) in model.GS for t in model.TT
    )
model.SCCConstraint = Constraint(rule=scc_rule)

# Constraint for PipeCC (Pipeline Capital Cost)
def pipecc_rule(model):
    return 0.001*model.PipeCC == 0.001*(sum(
        # First summation: Hydrogen pipeline cost
        model.dfc[t] * model.cccH[d1] * model.DistPipe[g, g1] * model.Yh[d1, (g, g1), t]
        for d1 in model.d1 if model.ord_d1[d1]==3
        for g in model.g
        for g1 in model.g
        if (g, g1) in model.Npipe and model.ord_g[g] < model.ord_g[g1]
        for t in model.TT
    ) + sum(
        # Second summation: Onshore CO2 cost
        model.dfc[t] * model.cccC_onshore[d2] * model.Dist[g, g1] * model.Yon[d2, (g, g1), t]
        for d2 in model.d2 if model.ord_d2[d2]==2
        for g in model.g
        for g1 in model.g
        if (g, g1) in model.N and model.ord_g[g] < model.ord_g[g1]
        for t in model.TT
    ) + sum(
        # Third summation: Offshore CO2 cost
        model.dfc[t] * model.cccC_offshore[d2] * model.DistRes[g, r] * model.Yoff[d2, (g, r), t]
        for d2 in model.d2 if model.ord_d2[d2]==2
        for g in model.g
        for r in model.r
        if (g, r) in model.GR
        for t in model.TT
    ) + sum(
        # Fourth summation: Storage cost
        model.dfc[t] * model.cccH[d1] * model.DistSt[g, sc] * model.Yst[d1, (g, sc), t]
        for d1 in model.d1 if model.ord_d1[d1]==3
        for g in model.g
        for sc in model.sc
        if (g, sc) in model.GS2
        for t in model.TT
    ))
model.PipeCCConstraint = Constraint(rule=pipecc_rule)

# Constraint for TCC (Total Capital Cost)
def tcc_rule(model):
    return model.TCC == 1000*model.PipeCC
model.TCCConstraint = Constraint( rule=tcc_rule)


# Constraint for POC
def poc_rule(model,k):
    return 0.001*model.POC[k] == 0.001*sum(
        model.dfo[t] * (
            model.pocostKF[p, t,k] * model.PCap[p] * model.NP[p, g, t] +
            sum(
                model.WF[c] * model.pocostKV[p, t,k] * model.theta[c,h] * model.Pr[p, g, t, c, h,k]
                for c in model.CC for h in model.h if (c,h) in model.CH
            )
        )
        for p in model.p for g in model.g for t in model.TT
    )
model.POCConstraint = Constraint(model.k, rule=poc_rule)

# Constraint for SOC
def soc_rule(model,k):
    return 0.001*model.SOC[k] == 0.001* sum(
        model.dfo[t] * (
            model.socostF[s] * model.SCap[s] * model.NS[s, g, t] +
            sum(
                model.WF[c] * model.socostV[s] * model.theta[c,h] * model.Qi[g, s, t, c, h,k]
                for c in model.CC for h in model.h if (c,h) in model.CH
            )
        )
        for s in model.s for g in model.g if (g, s) in model.GS for t in model.TT
    )
model.SOCConstraint = Constraint(model.k, rule=soc_rule)

# Constraint for PipeOC
def pipeoc_rule(model):
    return 0.001*model.PipeOC == 0.001*(sum(
        model.dfo[t] * model.deltaH * model.crf * model.cccH[d1] * model.DistPipe[g, g1] * model.AY[d1, (g, g1), t]
        for d1 in model.d1 if model.ord_d1[d1]==3
        for g in model.g 
        for g1 in model.g 
        if (g, g1) in model.Npipe and model.ord_g[g] < model.ord_g[g1]
        for t in model.TT
    ) + sum(
        model.dfo[t] * model.deltaC_onshore * model.crf * model.cccC_onshore[d2] * model.Dist[g, g1] * model.AYon[d2, (g, g1), t]
        for d2 in model.d2 if model.ord_d2[d2]==2
        for g in model.g 
        for g1 in model.g 
        if (g, g1) in model.N and model.ord_g[g] < model.ord_g[g1]
        for t in model.TT
    ) + sum(
        model.dfo[t] * model.deltaC_offshore * model.crf * model.cccC_offshore[d2] * model.DistRes[g, r] * model.AYoff[d2, (g, r), t]
        for d2 in model.d2 if model.ord_d2[d2]==2
        for g in model.g 
        for r in model.r if (g, r) in model.GR 
        for t in model.TT
    ) + sum(
        model.dfo[t] * model.deltaH * model.crf * model.cccH[d1] * model.DistSt[g, sc] * model.AYst[d1, (g, sc), t]
        for d1 in model.d1 if model.ord_d1[d1]==3
        for g in model.g 
        for sc in model.sc if (g, sc) in model.GS2 
        for t in model.TT
    ))
model.PipeOCConstraint = Constraint(rule=pipeoc_rule)

# Constraint for TOC
def toc_rule(model,k):
    return model.TOC[k] ==  1000 * model.PipeOC
model.TOCConstraint = Constraint(model.k, rule=toc_rule) 


# Constraint for CEC
def cec_rule(model,k):
    return model.CEC[k] == sum(
        model.WF[c] * model.dfo[t] * model.ct[t] * model.y_e[p, t] * model.theta[c,h] * model.Pr[p, g, t, c, h,k]
        for p in model.p for g in model.g for t in model.TT for c in model.CC for h in model.h if (c,h) in model.CH
    )
model.CECConstraint = Constraint(model.k, rule=cec_rule)

# Constraint for IIC
def iic_rule(model,k):
    return model.IIC[k] == sum(
        model.WF[c] * model.dfo[t] * model.pimp * model.theta[c,h] * model.IMP[g, t, c, h,k]
        for g in model.Gimp for t in model.TT for c in model.CC for h in model.h if (c,h) in model.CH
    )
model.IICConstraint = Constraint(model.k, rule=iic_rule)

# Constraint for ReC
def rec_rule(model,k):
    return 0.001*model.ReC[k] == 0.001*sum(
        model.dfc[t] * model.rccost[e, t,k]* model.InvR[e, g, t] +
        model.dfo[t] * model.rocost[e, t] * model.NR[e, g, t]
        for t in model.TT for e in model.e for g in model.g
    )
model.ReCConstraint = Constraint(model.k, rule=rec_rule)

# Constraint for GC
def gc_rule(model,k):
    return model.GC[k] == sum(
        model.dfo[t] * model.cgas[t,k] * model.Vgas[t,k]
        for t in model.TT
    )
model.GCConstraint = Constraint(model.k, rule=gc_rule)

# Constraint for BC
def bc_rule(model,k):
    return model.BC[k] == sum(
        model.dfo[t] * model.cbio[t] * model.Vbio[t,k]
        for t in model.TT
    )
model.BCConstraint = Constraint(model.k, rule=bc_rule)

def TC_rule(model,k):
    return model.TC[k]==(
        1000*model.PCC[k]+ 
        model.SCC+
        model.TCC+
        model.POC[k]+ 
        model.SOC[k]+ 
        +model.TOC[k]+
        model.CEC[k]+ 
        model.IIC[k]+ 
        model.ReC[k]+
        model.GC[k]+ 
        model.BC[k]+sum(model.WF[c] * model.dfo[t] * 5*model.pimp * model.theta[c,h]*model.slak1[g,t,c,h,k] 
                        for g in model.g for t in model.TT for c in model.CC for h in model.h if (c,h) in model.CH) 
        )

model.TCConstraint= Constraint(model.k, rule=TC_rule)

def objective_rule(model):
    return (sum(model.prob[k]*model.TC[k] for k in model.k)
    )   
                  
model.Total = Objective(rule=objective_rule, sense=minimize)



# %%
#======================================Problem Constraints========================================================*



#---------- FUELS CONSUMPTION --------------*
# %%  Fuel Constraint ----------------------------
def gas_cons_rule(model, t,k):
    if t in model.TT:
        return model.Vgas[t,k] == sum(
            model.WF[c] * model.theta[c,h] * model.Pr[p, g, t, c, h,k] / model.etaK[p, t,k]
            for p in model.p if model.ord_p[p] <= 2 for g in model.g for c in model.CC for h in model.h if (c,h) in model.CH
            
        )
    return Constraint.Skip

model.GasConsConstraint = Constraint(model.TT, model.k, rule=gas_cons_rule)

# Biomass consumption
def bio_cons_rule(model, t,k):
    if t in model.TT:
        return model.Vbio[t,k] == sum(
            model.WF[c] * model.theta[c,h] * model.Pr['BECCS', g, t, c, h,k] / model.etaK['BECCS', t,k]
            for g in model.g for c in model.CC for h in model.h if (c,h) in model.CH        )
    return Constraint.Skip

model.BioConsConstraint = Constraint(model.TT, model.k, rule=bio_cons_rule)



# Biomass Availability 

def biomass_availability_rule(model, g, t,k):
    if t in model.TT:  # Apply the constraint only for TT(t)
        return sum(
            0.001* model.WF[c] * model.theta[c,h] * model.Pr['BECCS', g, t, c, h,k] / model.etaK['BECCS', t,k]
            for c in model.CC for h in model.h if (c,h) in model.CH
        ) <= 0.001* model.BA[g,t,k]
    return Constraint.Skip

model.BiomassAvailabilityConstraint = Constraint(model.g, model.TT, model.k, rule=biomass_availability_rule)
# %%  RAMP UP/DOWN -------------------------------

# Ramp Up
def ramp_up_rule(model, p, g, c, h, t,k):
    if t in model.TT and c in model.CC and h in model.h and h > 1 and (c,h) in model.CH:
        return model.Pr[p, g, t, c, h,k] - model.Pr[p, g, t, c, h - 1,k] <=model.theta[c,h] * model.RU[p] * model.PCap[p] * model.NP[p, g, t]
    return Constraint.Skip

model.RampUpConstraint = Constraint(model.p, model.g, model.CC, model.h, model.TT,model.k, rule=ramp_up_rule)


# Ramp Down
def ramp_down_rule(model, p, g, c, h, t,k):
    if t in model.TT and c in model.CC and h in model.h and h > 1 and (c,h) in model.CH:
        return model.Pr[p, g, t, c, h - 1,k] - model.Pr[p, g, t, c, h,k] <= model.theta[c,h] * model.RD[p] * model.PCap[p] * model.NP[p, g, t]
    return Constraint.Skip

model.RampDownConstraint = Constraint(model.p, model.g, model.CC, model.h, model.TT, model.k, rule=ramp_down_rule)
# %%  Peoduction Limit ---------------------------
# Production Limit

def p_capacity2_rule(model, p, g, t, c, h,k):
    if t in model.TT and c in model.CC and h in model.h and (c,h) in model.CH:
        return model.Pr[p, g, t, c, h,k] <= model.PCap[p] * model.pcap_max[p] * model.NP[p, g, t]
    return Constraint.Skip

model.PCapacity2Constraint = Constraint(model.p, model.g, model.TT, model.CC, model.h,model.k,  rule=p_capacity2_rule)


def p_availability_rule(model, p, g, t):
     if t in model.TT:
      return model.NP[p, g, t] == (
            model.NP[p, g, t-(model.dur/5)] if t>model.y1 else 0 
        )+( 
        model.np0[p, g] if t==model.y1 else 0
        )+ model.InvP[p, g, t] 
     return Constraint.Skip

model.PAvailability = Constraint(model.p, model.g, model.TT, rule=p_availability_rule)
 
#---------- STORAGE CONSTRAINTS --------------*
# %%  Storage Limit ------------------------------
# Storage 


def sinventory2_rule(model, s, g, t, c, h,k):
    if (g, s) in model.GS and t in model.TT and c in model.CC and h in model.h and (c,h) in model.CH:
        return model.St[s, g, t, c, h,k] == (
            (model.St[s, g, t, c, h - 1,k] if h > 1 else model.st0[s, g]) 
            + model.theta[c,h] * (model.Qi[g, s, t, c, h,k] - model.Qr[s, g, t, c, h,k])
        )
    return Constraint.Skip
model.SInventory2 = Constraint(model.s, model.g, model.TT, model.CC, model.h, model.k, rule=sinventory2_rule)


# Maximum injection rate
def max_inj_rule(model, s,g,t, c, h,k):
    if (g, s) in model.GS and t in model.TT and c in model.CC and h in model.h and (c,h) in model.CH:
        return model.Qi[(g, s), t, c, h,k] <= model.QImax[s] * model.NS[s, g, t]
    return Constraint.Skip

model.MaxInjConstraint = Constraint( model.GS, model.TT, model.CC, model.h, model.k, rule=max_inj_rule)


# Maximum retrieval rate
def max_retr_rule(model, s, g, t, c, h,k):
    if (g, s) in model.GS and t in model.TT and c in model.CC and h in model.h and (c,h) in model.CH:
        return model.Qr[s, g, t, c, h,k] <= model.QRmax[s] * model.NS[s, g, t]
    return Constraint.Skip
model.MaxRetrConstraint = Constraint(model.s,model.g, model.TT, model.CC, model.h, model.k, rule=max_retr_rule)

# Underground storage capacity
def s_capacity_u_rule(model, sc, g, t, c, h):
    if (g,sc) in model.GS2 and t in model.TT and c in model.CC and h in model.h :
        return model.InvS[sc, g, t] <= sum(model.Yst[d1, g, sc, t] for d1 in model.d1 if model.ord_d1[d1]==3)
    return Constraint.Skip

model.SCapacityUConstraint = Constraint(model.GS2, model.TT, model.CC, model.h, rule=s_capacity_u_rule)


# Storage capacity constraints
def s_capacity1_rule(model, s, g, t, c, h,k):
    if (g, s) in model.GS and t in model.TT and c in model.CC and h in model.HH and (c,h) in model.CH:
        return model.BS[s,g,t,c,k]+ model.St[s, g, t, c, h,k] >= model.SCap[s] * model.scap_min[s] * model.NS[s, g, t]
    return Constraint.Skip

model.SCapacity1Constraint = Constraint(model.s,model.g, model.TT, model.CC, model.HH, model.k, rule=s_capacity1_rule)

def s_capacity2_rule(model, s, g, t, c, h,k):
    if (g, s) in model.GS and t in model.TT and c in model.CC and h in model.h and (c,h) in model.CH:
        return model.BS[s,g,t,c,k]+ model.St[s, g, t, c, h,k] <= model.SCap[s] * model.scap_max[s] * model.NS[s, g, t]
    return Constraint.Skip

model.SCapacity2Constraint = Constraint(model.s,model.g, model.TT, model.CC, model.h, model.k, rule=s_capacity2_rule)



# Storage facility availability

def savailability_rule(model, s, g, t):
    if (g, s) in model.GS and t in model.TT:
        return model.NS[s, g, t] == (
            (model.NS[s, g, t -(model.dur/5)] if t > model.y1 else model.ns0[s, g])
            + model.InvS[s, g, t]
        )
    return Constraint.Skip

model.SAvailability = Constraint(model.s, model.g, model.TT, rule=savailability_rule)


def BSeq1_rule(model,s,g,t,c,k):
    if (g,s) in model.GS:
        return model.BS[s,g,t,1,k]==model.BS[s,g,t,5,k]
    return Constraint.Skip
model.BSeq1Constraint=Constraint(model.s,model.g,model.t,model.CC, model.k, rule=BSeq1_rule)

def BSeq2_rule(model,s,g,t,c,k):
    if (g,s) in model.GS and model.ord_c[c]>=3:
        return model.BS[s,g,t,c,k]==model.BS[s,g,t,c-1,k]+model.WF[c-1]*(sum(model.St[s,g,t,c-1,h,k]-model.St[s,g,t,c-1,1,k] for h in model.h if model.ord_h[h]==model.hf))
    return Constraint.Skip
model.BSeq2Constraint=Constraint(model.s,model.g,model.t,model.CC, model.k, rule=BSeq2_rule)

def BSeq3_rule(model,s,g,t,c,k):
    if (g,s) in model.GS:
        return model.BS[s,g,t,2,k]==model.BS[s,g,t,5,k]+model.WF[5]*(sum(model.St[s,g,t,5,h,k]-model.St[s,g,t,5,1,k] for h in model.h if model.ord_h[h]==model.hf))+model.St[s,g,t,1,24,k]-model.St[s,g,t,1,1,k]
    return Constraint.Skip
model.BSeq3Constraint=Constraint(model.s,model.g,model.t,model.CC, model.k, rule=BSeq3_rule)
#---------- RENEWABLES CONSTRAINTS --------------*
# %%  RENEWABLES CONSTRAINTS ---------------------
# Electricity production for electrolysis
def elec_prod_rule(model, g, t, c, h,k):
    if (c,h) in model.CH:
        return model.Pr['WE', g, t, c, h,k] == model.etaK['WE', t,k] * (
            sum(model.Pre[e, g, t, c, h,k] for e in model.e) - model.CL[g, t, c, h,k]
        )
    return Constraint.Skip

model.ElecProdConstraint = Constraint(model.g, model.TT, model.CC, model.h, model.k, rule=elec_prod_rule)


# Renewables availability
def renew_av_rule(model, e, g, t, c, h,k):
    if t in model.TT and c in model.CC and (c,h) in model.CH:
        return model.Pre[e, g, t, c, h,k] ==  0.7*model.AV[c, h,e,g, k] * model.NR[e, g, t]
    return Constraint.Skip
model.RenewAvConstraint = Constraint(model.e, model.g, model.TT, model.CC, model.HH,model.k, rule=renew_av_rule)



def renew_cap_rule(model, e, g, t):
    if t in model.TT:
            return model.NR[e, g, t] == (model.NR[e, g, t -(model.dur/5)] if t >model.y1 else 0)+ model.InvR[e, g, t]
    return Constraint.Skip

model.RenewCapConstraint = Constraint(model.e, model.g, model.TT, rule=renew_cap_rule)


def land_availability_rule(model, e, g, t):
    if t in model.TT:
        return 0.001*model.NR[e, g, t] <= 0.001*model.landAV[e, g]
    return Constraint.Skip

model.LandAvailabilityConstraint = Constraint(model.e, model.g, model.TT, rule=land_availability_rule)




# %%  Hydrogen and CO2 blance --------------------
# HA2 Monolithic Version


def flow_balance_rule(model, g, t, c, h,k):
    if (c,h) in model.CH:
        return (
            sum(model.Pr[p, g, t, c, h,k] for p in model.p) +
            sum(model.Q['Pipe', g1, g, t, c, h,k] for g1 in model.g if (g1, g) in model.Npipe) +
            (model.IMP[g, t, c, h,k] if g in model.Gimp else 0) +
            sum(model.Qr[s, g, t, c, h,k] for s in model.s if (g, s) in model.GS)
            ==
            sum(model.Q['Pipe', g, g1, t, c, h,k] for g1 in model.g if (g, g1) in model.Npipe) +
            sum(model.Qi[g, s, t, c, h,k] for s in model.s if (g, s) in model.GS) +
            model.dem[g, t, c, h,k]
        )
    return Constraint.Skip
    

model.FlowBalance = Constraint(model.g, model.TT, model.CC, model.h,model.k, rule=flow_balance_rule)



# Co2 Balanace
def co2_mass_balance_rule(model, g, t, c, h,k):
    if t in model.TT and c in model.CC and h in model.h and (c,h) in model.CH:
        return (
            sum(model.Qon[g1, g, t, c, h,k] for g1 in model.g if (g1, g) in model.N) +
            sum(model.y_c[p, t] * model.Pr[p, g, t, c, h,k] for p in model.p)
         == 
            sum(model.Qon[g, g1, t, c, h,k] for g1 in model.g if (g, g1) in model.N) +
            sum(model.Qoff[g, r, t, c, h,k] for r in model.r if (g, r) in model.GR)
        )
    return Constraint.Skip

model.CO2MassBalanceConstraint = Constraint(model.g, model.TT, model.CC, model.h,model.k, rule=co2_mass_balance_rule)




#---------- RESERVIORS Constraints --------------*
# %%  RESERVIORS Constraints ---------------------
# Inventory
def res_inventory_rule(model, r, t,k):
    if t in model.TT: 
      return model.RI[r, t,k] == (model.RI[r, t - (model.dur/5),k] if t>model.y1 else  model.ri0[r] / 1000)+ model.dur * sum(
        model.WF[c] * model.theta[c,h] * model.Qoff[(g, r), t, c, h,k] for g in model.g if 
        (g, r) in model.GR  for c in model.CC for h in model.h
         ) / 1000
   

model.ResInventoryConstraint = Constraint(model.r, model.TT,model.k, rule=res_inventory_rule)


# %%  Hydrogen Import Limit ----------------------
# Import limit
def imp_limit_rule(model, t, c, h,k):
    if t in model.TT and c in model.CC and h in model.h and (c,h) in model.CH:
        return sum(model.IMP[g, t, c, h,k] for g in model.Gimp) <= 3*model.iota*sum(model.dem[g, t, c, h,k] for g in model.g)
    return Constraint.Skip
model.ImpLimitConstraint = Constraint(model.TT, model.CC, model.h, model.k, rule=imp_limit_rule)


#---------- EMISSION --------------*
# %%  Emission Target Limit ----------------------
# Emissions target
def emissions_rule(model, t,k):
    if t in model.TT:
        return model.em[t,k] == sum(
            model.WF[c] * model.y_e[p, t] * model.theta[c,h] * model.Pr[p, g, t, c, h,k]
            for p in model.p
            for g in model.g
            for c in model.CC
            for h in model.h if (c,h) in model.CH
        )
    return Constraint.Skip
model.EmissionConstraint = Constraint(model.TT,model.k, rule=emissions_rule)

# Emissions target equation
def em_target_rule(model, t,k):
  
        return model.em[6,k] <= 0
   
model.EmTargeteqConstraint = Constraint(model.TT, model.k, rule=em_target_rule)
# %%  Tighting -----------------------------------
# Tight Gas
'''
# TightGas Constraint
def tight_gas_rule(model, t):
    if t in model.TT:
        return sum(
            model.InvP[p, g, t] for g in model.g for p in model.p if model.ord_p[p] <= 2
        ) <= 40
    return Constraint.Skip

model.TightGasConstraint = Constraint(model.TT, rule=tight_gas_rule)


# TightBio Constraint
def tight_bio_rule(model, p, t):
    if model.ord_p[p] == 3 and t in model.TT:
        return sum(
            model.InvP[p, g, t] for g in model.g
        ) <= 30
    return Constraint.Skip

model.TightBioConstraint = Constraint(model.p, model.TT, rule=tight_bio_rule)


# TightInvWE Constraint
def tight_inv_we_rule(model, p, t):
    if model.ord_p[p] == 4 and t in model.TT:
        return sum(
            model.InvP[p, g, t] for g in model.g
        ) <= 200
    return Constraint.Skip

model.TightInvWEConstraint = Constraint(model.p, model.TT, rule=tight_inv_we_rule)


# TightInvStorage Constraint
def tight_inv_storage_rule(model, s, t):
    if model.ord_s[s] > 4 and t in model.TT:
        return sum(
            model.InvS[s, g, t] for g in model.g
        ) <= 250
    return Constraint.Skip

model.TightInvStorageConstraint = Constraint(model.s, model.TT, rule=tight_inv_storage_rule)
'''

# %%  PIPELINE CONSTRAINTS------------------------
#------Hydrogen Pipeline Limit ------

# Maximum flowrate for pipelines

def h2pipe_max_rule(model, g, g1, t, c, h,k):
    if (g, g1) in model.Npipe and (c,h) in model.CH:
        return model.Q['Pipe', (g, g1), t, c, h,k] <= sum(model.qHmax[d1] * (
            (model.AY[d1, (g, g1), t] if model.ord_g[g] < model.ord_g[g1] else 0)+
            (model.AY[d1, (g1, g), t] if model.ord_g[g1] < model.ord_g[g] else 0) 
            )
            for d1 in model.d1 if model.ord_d1[d1]==3)
    return Constraint.Skip

model.H2PipeMax = Constraint(model.Npipe, model.TT, model.CC, model.h, model.k, rule=h2pipe_max_rule)



def onshorepipe_max_rule(model, g, g1, t, c, h,k):
    if (g, g1) in model.N and (c,h) in model.CH:
        return 0.001*model.Qon[g, g1, t, c, h,k] <= 0.001*sum(model.qCmax[d2] * (
                (model.AYon[d2, g, g1, t] if model.ord_g[g] < model.ord_g[g1] else 0) +
                (model.AYon[d2, g1, g, t] if model.ord_g[g1] < model.ord_g[g] else 0)
            )
            for d2 in model.d2 if model.ord_d2[d2]==2
        )
    return Constraint.Skip

model.OnshorePipeMax = Constraint(model.N, model.TT, model.CC, model.h,model.k, rule=onshorepipe_max_rule)




def offshorepipe_max_rule(model, g, r, t, c, h,k):
    if (g, r) in model.GR and (c,h) in model.CH:
        return 0.001*model.Qoff[(g, r), t, c, h,k]  <= 0.001*sum(model.qCmax[d2] * model.AYoff[d2, (g, r), t] for d2 in model.d2 if model.ord_d2[d2]==2)
    return Constraint.Skip
model.OffshorePipeMax = Constraint(model.GR, model.TT, model.CC, model.h, model.k, rule=offshorepipe_max_rule)


# Availability of pipelines
def H2PAvailability_rule(model, d1, g, g1, t):
    if (g,g1) in model.Npipe and t in model.TT and model.ord_g[g] < model.ord_g[g1] and  model.ord_d1[d1]==3:
        return model.AY[d1, g, g1, t] == (
            model.AY[d1, g, g1, t - (model.dur/5)] if t > model.y1 else 0
        ) + (model.ayHR0[d1, g, g1] if t == model.y1 else 0) + model.Yh[d1, g, g1, t]
    return Constraint.Skip

model.H2PAvailability = Constraint(model.d1, model.Npipe, model.TT, rule=H2PAvailability_rule)


def onp_availability_rule_simple(model, d2, g, g1, t):
       if (g, g1) in model.N and model.ord_g[g] < model.ord_g[g1] and model.ord_d2[d2]==2:
            return model.AYon[d2, (g, g1), t] == (
                model.AYon[d2, (g, g1), t - (model.dur/5)] if t>model.y1 else 0
                )+ (model.ayC0[d2, (g, g1)] if t ==model.y1 else 0 ) + model.Yon[d2, (g, g1), t]
       return Constraint.Skip
model.OnPAvailability = Constraint(model.d2, model.N, model.TT, rule=onp_availability_rule_simple)


def offp_availability_rule(model, d2, g, r, t):
     if (g, r) in model.GR and t in model.TT and model.ord_d2[d2]==2: 
         return model.AYoff[d2, (g, r), t] == (
             model.AYoff[d2, (g, r), t-(model.dur/5)] if t> model.y1 else 0
             )+ (model.aeC0[r] if t==model.y1 else 0) + model.Yoff[d2, (g, r), t]
     return Constraint.Skip

model.OffPAvailability = Constraint(model.d2, model.GR, model.TT, rule=offp_availability_rule)

def pipest_availability_rule(model, d1, g, sc, t):
      if (g, sc) in model.GS2 and t in model.TT and model.ord_d1[d1]==3: 
           return model.AYst[d1, (g, sc), t] == (
               model.AYst[d1, g, sc, t-(model.dur/5)]  if t>model.y1 else 0)+ model.Yst[d1, (g, sc), t]
      return Constraint.Skip
model.PipeStAvailability = Constraint(model.d1, model.GS2, model.TT, rule=pipest_availability_rule)


# One diameter size
def h2pipe_rule(model, g, g1, t):
     if (g,g1) in model.Npipe and model.ord_g[g] < model.ord_g[g1] and t in model.TT:   
       return sum(model.AY[d1, (g, g1), t] for d1 in model.d1) <= 1
     return Constraint.Skip
model.H2Pipe = Constraint(model.Npipe, model.TT, rule=h2pipe_rule)



def onpipe_rule(model, g, g1, t):
    if (g, g1) in model.N and model.ord_g[g] < model.ord_g[g1] and t in model.TT:
        return sum(model.AYon[d2, g, g1, t] for d2 in model.d2) <= 1
    return Constraint.Skip

model.OnPipeConstraint = Constraint(model.N, model.TT, rule=onpipe_rule)



def offpipe_rule(model, g, r, t):
     if (g,r) in model.GR and t in model.TT:  
        return sum(model.AYoff[d2, (g, r), t] for d2 in model.d2) <= 1
        return Constraint.Skip
model.OffPipe = Constraint(model.GR, model.TT, rule=offpipe_rule)

def stpipe_rule(model, g, sc, t):
    if (g,sc) in model.GS2:
        return sum(model.AYst[d1, (g, sc), t] for d1 in model.d1) <= 1
    return Constraint.Skip
model.StPipe = Constraint(model.GS2, model.TT, rule=stpipe_rule)
# %%  Solving the model

import tkinter as tk
from pyomo.environ import SolverFactory, SolverManagerFactory
import os


def center_window(win, width=400, height=300):
    win.update_idletasks()
    screen_width = win.winfo_screenwidth()
    screen_height = win.winfo_screenheight()
    x = (screen_width - width) // 2
    y = (screen_height - height) // 2
    win.geometry(f"{width}x{height}+{x}+{y}")


def get_solver_mode():
    def submit():
        nonlocal selected_mode
        selected_mode = var.get()
        dialog.destroy()
        root.quit()

    root = tk.Tk()
    root.withdraw()

    selected_mode = None

    dialog = tk.Toplevel(root)
    dialog.title("Select Solver Mode")
    dialog.configure(bg="#2c3e50")
    center_window(dialog, 400, 250)

    tk.Label(
        dialog,
        text="Choose Solver Mode:",
        font=("Arial", 12, "bold"),
        fg="white",
        bg="#2c3e50"
    ).pack(pady=10)

    var = tk.StringVar(value="local")

    modes = [
        ("Local Solver", "local"),
        ("NEOS Solver", "neos")
    ]

    for text, value in modes:
        tk.Radiobutton(
            dialog,
            text=text,
            variable=var,
            value=value,
            font=("Arial", 11, "bold"),
            fg="black",
            bg="#ecf0f1",
            anchor="w"
        ).pack(fill="x", padx=20, pady=5)

    tk.Button(
        dialog,
        text="Confirm",
        font=("Arial", 11, "bold"),
        fg="white",
        bg="#2980b9",
        width=15,
        height=2,
        command=submit
    ).pack(pady=15)

    dialog.grab_set()
    root.mainloop()

    return selected_mode


def get_solver():
    def submit():
        nonlocal selected_solver
        selected_solver = var.get()
        dialog.destroy()
        root.quit()

    root = tk.Tk()
    root.withdraw()

    selected_solver = None

    dialog = tk.Toplevel(root)
    dialog.title("Select Local Solver")
    dialog.configure(bg="#2c3e50")
    center_window(dialog, 400, 300)

    tk.Label(
        dialog,
        text="Choose a solver:",
        font=("Arial", 12, "bold"),
        fg="white",
        bg="#2c3e50"
    ).pack(pady=10)

    var = tk.StringVar(value="gurobi")

    solvers = [
        ("Gurobi", "gurobi"),
        ("CPLEX", "cplex"),
        ("GLPK", "glpk"),
        ("HiGHS", "highs")
    ]

    for text, value in solvers:
        tk.Radiobutton(
            dialog,
            text=text,
            variable=var,
            value=value,
            font=("Arial", 11, "bold"),
            fg="black",
            bg="#ecf0f1",
            anchor="w"
        ).pack(fill="x", padx=20, pady=5)

    tk.Button(
        dialog,
        text="Confirm",
        font=("Arial", 11, "bold"),
        fg="white",
        bg="#2980b9",
        width=15,
        height=2,
        command=submit
    ).pack(pady=15)

    dialog.grab_set()
    root.mainloop()

    return selected_solver


def get_solver_settings():
    time_limit = 3600
    mip_gap = 0.05

    def submit_settings():
        nonlocal time_limit, mip_gap

        try:
            time_limit = int(time_limit_entry.get())
            mip_gap = float(mip_gap_entry.get())
        except ValueError:
            error_label.config(
                text="Invalid input! Enter numbers only.",
                fg="red"
            )
            return

        settings_dialog.destroy()

    settings_dialog = tk.Toplevel()
    settings_dialog.title("Solver Settings")
    settings_dialog.configure(bg="#34495e")
    center_window(settings_dialog, 400, 250)

    tk.Label(
        settings_dialog,
        text="Enter Solver Settings",
        font=("Arial", 14, "bold"),
        fg="white",
        bg="#34495e"
    ).pack(pady=10)

    tk.Label(
        settings_dialog,
        text="Time Limit (seconds):",
        font=("Arial", 12),
        fg="white",
        bg="#34495e"
    ).pack()

    time_limit_entry = tk.Entry(settings_dialog, font=("Arial", 12))
    time_limit_entry.insert(0, str(time_limit))
    time_limit_entry.pack(pady=5)

    tk.Label(
        settings_dialog,
        text="MIP Gap (e.g. 0.05):",
        font=("Arial", 12),
        fg="white",
        bg="#34495e"
    ).pack()

    mip_gap_entry = tk.Entry(settings_dialog, font=("Arial", 12))
    mip_gap_entry.insert(0, str(mip_gap))
    mip_gap_entry.pack(pady=5)

    error_label = tk.Label(
        settings_dialog,
        text="",
        font=("Arial", 10, "bold"),
        bg="#34495e"
    )
    error_label.pack()

    tk.Button(
        settings_dialog,
        text="Apply",
        font=("Arial", 12, "bold"),
        fg="white",
        bg="#27ae60",
        width=15,
        height=2,
        command=submit_settings
    ).pack(pady=15)

    settings_dialog.grab_set()
    settings_dialog.wait_window()

    return time_limit, mip_gap


def get_neos_email():
    email = "m.hemmati@ucl.ac.uk"

    def submit_email():
        nonlocal email
        email = email_entry.get().strip()

        if not email:
            error_label.config(
                text="Email cannot be empty.",
                fg="red"
            )
            return

        dialog.destroy()

    dialog = tk.Toplevel()
    dialog.title("NEOS Settings")
    dialog.configure(bg="#34495e")
    center_window(dialog, 400, 220)

    tk.Label(
        dialog,
        text="Enter NEOS Email",
        font=("Arial", 14, "bold"),
        fg="white",
        bg="#34495e"
    ).pack(pady=10)

    email_entry = tk.Entry(dialog, font=("Arial", 12), width=35)
    email_entry.insert(0, email)
    email_entry.pack(pady=10)

    error_label = tk.Label(
        dialog,
        text="",
        font=("Arial", 10, "bold"),
        bg="#34495e"
    )
    error_label.pack()

    tk.Button(
        dialog,
        text="Confirm",
        font=("Arial", 12, "bold"),
        fg="white",
        bg="#2980b9",
        width=15,
        height=2,
        command=submit_email
    ).pack(pady=15)

    dialog.grab_set()
    dialog.wait_window()

    return email


# =========================
# Main Flow
# =========================

solver_mode = get_solver_mode()

if solver_mode == "local":
    solver_name = get_solver()

    if not solver_name:
        print("No local solver selected.")
        exit()

    print(f"Selected local solver: {solver_name}")

    time_limit, mip_gap = get_solver_settings()

    opt = SolverFactory(solver_name)
    opt.options["Threads"] = 36
    opt.options["Presolve"] = 2
    opt.options["MIPGap"] = mip_gap
    opt.options["TimeLimit"] = time_limit
    opt.options["Heuristics"] = 0.1

    results = opt.solve(model, tee=True)


elif solver_mode == "neos":
    email = get_neos_email()

    print(f"Using NEOS email: {email}")

    os.environ["NEOS_EMAIL"] = email

    solver_manager = SolverManagerFactory("neos")

    results = solver_manager.solve(
        model,
        opt="cplex",
        tee=True,
        keepfiles=False,
        load_solutions=True,
        timelimit=3600
    )

else:
    print("No solver mode selected.")
    exit()


print("Solver finished successfully.")

# %%
import tkinter as tk
from tkinter import ttk
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

# ===============================
# Step 1: Collect and format TC[k]
# ===============================

costs = []
for k in model.k:
    val = model.TC[k].value / 1e9  # Convert to billions
    costs.append((f"TC[{k}]", val))

# ===============================
# Step 2: Build tkinter window
# ===============================

root = tk.Tk()
root.title("Cost Breakdown (£b)")

window_width = 600
window_height = 800
screen_width = root.winfo_screenwidth()
screen_height = root.winfo_screenheight()
position_top = int(screen_height / 2 - window_height / 2)
position_left = int(screen_width / 2 - window_width / 2)

root.geometry(f"{window_width}x{window_height}+{position_left}+{position_top}")
root.configure(bg="#f4f4f4")
root.attributes("-topmost", True)

title = tk.Label(root, text="📊 Cost Breakdown (£b)", font=("Helvetica", 18, "bold"), bg="white", fg="blue")
title.pack(pady=20)

tree = ttk.Treeview(root, columns=("Cost Type", "Amount"), show="headings", height=12)
tree.heading("Cost Type", text="Cost Type", anchor="w")
tree.heading("Amount", text="Amount (£b)", anchor="w")

tree.column("Cost Type", width=200)
tree.column("Amount", width=150, anchor="e")

for i, (name, value) in enumerate(costs):
    tag = "even" if i % 2 == 0 else "odd"
    tree.insert("", "end", values=(name, f"{value:.3f}"), tags=(tag,))

tree.tag_configure("even", background="#f9f9f9")
tree.tag_configure("odd", background="#e0e0e0")
tree.pack(pady=10)

# ===============================
# Step 3: Draw the bar chart
# ===============================

def plot_bar_chart():
    labels = [name for name, _ in costs]
    values = [value for _, value in costs]

    

    fig, ax = plt.subplots(figsize=(8, 8))
    ax.bar(labels, values)

    ax.set_title("Cost per sceanrio (£b)", fontsize=16, color='darkblue')
    ax.set_xlabel("Scenario", fontsize=8)
    ax.set_ylabel("Amount (£b)", fontsize=12)
    ax.set_xticklabels(labels, rotation=45, ha="right", fontsize=10, color="black")

    canvas = FigureCanvasTkAgg(fig, master=root)
    canvas.draw()
    canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

plot_bar_chart()

root.mainloop()
